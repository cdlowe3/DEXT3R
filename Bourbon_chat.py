import openpyxl

def create_bourbon_collection():
  """Creates a list of bourbon dictionaries."""
  bourbons = [
    {'name': 'Pappy Van Winkle 23 Year', 'distillery': 'Willett Distillery', 'age': 23, 'proof': 114.2, 'price': 300, 'rating': 5, 'notes': 'Complex, rich, caramel'},
    # Add more bourbons here
  ]
  return bourbons

def save_to_excel(bourbons, filename='bourbon_collection.xlsx'):
  """Saves the bourbon collection to an Excel file."""
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  sheet.title = 'Bourbon Collection'

  # Write headers
  headers = ['Name', 'Category', 'Type', 'Supplier', 'UPC Code', 'Distillery', 'Age', 'Proof', 'Price', 'Rating', 'Notes', 'Website', 'Volume', 'Quantity']
  for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)

  # Write bourbon data
  for row, bourbon in enumerate(bourbons, start=2):
    for col, key in enumerate(headers, start=1):
      sheet.cell(row=row, column=col, value=bourbon.get(key, ''))

  workbook.save(filename)

if __name__ == '__main__':
  bourbon_collection = create_bourbon_collection()
  save_to_excel(bourbon_collection)



